#!/usr/bin/env python3
"""
GLM 智能体（Agent）批量测评脚本
支持：纯文本输入 / 图片 / 文件 / 音频 / 视频 批量测试
API 文档：https://zhipu-ai.feishu.cn/wiki/Wsr1wXmHXicO3AkdZPVcBANbnvb
"""

import argparse
import json
import os
import re
import sys
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests

# ── 常量 ─────────────────────────────────────────────────────────────────────

BASE_URL = "https://open.bigmodel.cn/api/llm-application/open"

IMAGE_EXTS = {".jpg", ".jpeg", ".png"}
FILE_EXTS  = {".pdf", ".doc", ".docx", ".ppt", ".pptx", ".txt", ".md",
              ".xlsx", ".xls", ".csv",
              ".mp3", ".m4a", ".wav", ".flac", ".ogg",
              ".mp4", ".mov", ".avi", ".mkv", ".webm"}

# file_type → MIME 映射（上传时用）
MIME_MAP = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc":  "application/msword",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.ms-excel",
    ".csv":  "text/csv",
    ".txt":  "text/plain",
    ".md":   "text/plain",
    ".mp3":  "audio/mpeg",
    ".m4a":  "audio/mp4",
    ".wav":  "audio/wav",
    ".flac": "audio/flac",
    ".ogg":  "audio/ogg",
    ".mp4":  "video/mp4",
    ".mov":  "video/quicktime",
    ".avi":  "video/x-msvideo",
    ".mkv":  "video/x-matroska",
    ".webm": "video/webm",
}

_excel_lock = threading.Lock()

# ── API 调用层 ────────────────────────────────────────────────────────────────

def _headers(api_key: str) -> dict:
    return {"Authorization": f"Bearer {api_key}"}


def list_variables(app_id: str, api_key: str) -> list:
    """查询智能体输入变量（获取 upload_unit_id 及变量名）"""
    resp = requests.get(
        f"{BASE_URL}/v2/application/{app_id}/variables",
        headers=_headers(api_key),
        timeout=15,
    )
    return resp.json().get("data", [])


def upload_files(app_id: str, api_key: str,
                 paths: list, upload_unit_id: str,
                 file_type: str = "4",
                 max_retries: int = 3) -> list[str]:
    """上传一批文件，返回成功的 file_id 列表"""
    url = f"{BASE_URL}/v2/application/file_upload"
    file_ids = []
    for path in paths:
        ext  = Path(path).suffix.lower()
        mime = MIME_MAP.get(ext, "application/octet-stream")
        fid  = None
        for attempt in range(1, max_retries + 1):
            try:
                with open(path, "rb") as f:
                    resp = requests.post(
                        url,
                        headers=_headers(api_key),
                        data={"app_id": app_id,
                              "upload_unit_id": upload_unit_id,
                              "file_type": file_type},
                        files={"files": (Path(path).name, f, mime)},
                        timeout=60,
                    )
                d = resp.json()
                if d.get("code") == 200:
                    items = d.get("data", {}).get("successInfo", [])
                    if items:
                        fid = items[0]["fileId"]
                        break
                    fail = d.get("data", {}).get("failInfo", [])
                    if fail:
                        print(f"  [上传失败] {Path(path).name}: "
                              f"{fail[0].get('failReason','未知')}", file=sys.stderr)
                        break
                else:
                    print(f"  [上传错误] code={d.get('code')} "
                          f"msg={d.get('message')}", file=sys.stderr)
            except Exception as e:
                print(f"  [上传异常] 第{attempt}次: {e}", file=sys.stderr)
            time.sleep(2 * attempt)
        if fid:
            file_ids.append(fid)
    return file_ids


def wait_for_files(app_id: str, api_key: str,
                   file_ids: list, max_wait: int = 60) -> bool:
    """轮询文件解析状态，code=1 或超时后返回"""
    if not file_ids:
        return False
    url     = f"{BASE_URL}/v2/application/file_stat"
    elapsed = 0
    while elapsed < max_wait:
        try:
            resp  = requests.post(url, headers={**_headers(api_key),
                                                "Content-Type": "application/json"},
                                  json={"app_id": app_id, "file_ids": file_ids},
                                  timeout=15)
            items = resp.json().get("data") or []
            if not items:
                # 图片上传后 file_stat 可能返回空列表，视为就绪
                return True
            if all(i.get("code") == 1 for i in items):
                return True
            # 非 0/1 的 code 说明出错，不再等待
            if any(i.get("code") not in (0, 1) for i in items):
                return True
        except Exception:
            pass
        time.sleep(2)
        elapsed += 2
    return True


def invoke_agent(app_id: str, api_key: str,
                 content_items: list,
                 max_retries: int = 3) -> Optional[str]:
    """
    调用推理接口（同步模式）。
    content_items: list of {"type":..., "value":..., "key":...}
    返回 msg 字符串，失败返回 None。
    """
    url = f"{BASE_URL}/v3/application/invoke"
    payload = {
        "app_id": app_id,
        "stream": False,
        "messages": [{"content": content_items}],
    }
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(
                url,
                headers={**_headers(api_key), "Content-Type": "application/json"},
                json=payload,
                timeout=120,
            )
            data = resp.json()
            choices = data.get("choices", [])
            if not choices:
                # 顶层 error_msg 场景
                err = data.get("error_msg") or data.get("message", "")
                if attempt < max_retries:
                    time.sleep(5 * attempt)
                    continue
                return f"[API错误] {err}"
            msg_obj = choices[0].get("messages", {}).get("content", {})
            msg = msg_obj.get("msg")
            if isinstance(msg, dict):
                return json.dumps(msg, ensure_ascii=False)
            if isinstance(msg, str):
                if "IllegalArgumentException" in msg and attempt < max_retries:
                    time.sleep(5 * attempt)
                    continue
                return msg
        except Exception as e:
            if attempt == max_retries:
                return f"[调用异常] {e}"
            time.sleep(3 * attempt)
    return None

# ── 任务执行 ──────────────────────────────────────────────────────────────────

def process_text_task(task: dict, args) -> dict:
    """纯文本任务"""
    start = time.time()
    result = {
        "index": task["index"],
        "input": task["prompt"],
        "input_files": [],
        "output": "",
        "success": False,
        "status": "success",
        "elapsed_ms": 0,
    }
    content = [{"type": "input", "value": task["prompt"], "key": args.input_key}]
    raw = invoke_agent(args.app_id, args.api_key, content)
    result["elapsed_ms"] = round((time.time() - start) * 1000)
    if raw and "调用异常" not in raw and "API错误" not in raw \
              and "IllegalArgumentException" not in raw:
        result["output"]  = raw
        result["success"] = True
    else:
        result["status"] = "invoke_failed"
        result["output"] = raw or ""
    return result


def process_file_task(task: dict, args) -> dict:
    """文件/图片/视频/音频任务"""
    start = time.time()
    paths = task["paths"]
    result = {
        "index": task["index"],
        "input": "",
        "input_files": [Path(p).name for p in paths],
        "output": "",
        "success": False,
        "status": "success",
        "elapsed_ms": 0,
    }

    # 上传
    file_ids = upload_files(args.app_id, args.api_key, paths,
                            args.upload_unit_id, str(args.file_type))
    if len(file_ids) < len(paths):
        result["status"] = f"upload_failed({len(file_ids)}/{len(paths)})"
        result["elapsed_ms"] = round((time.time() - start) * 1000)
        return result

    # 等待解析
    wait_for_files(args.app_id, args.api_key, file_ids)

    # 构造 content
    ext = Path(paths[0]).suffix.lower()
    if ext in IMAGE_EXTS:
        ctype = "upload_image"
    elif ext in {".mp4", ".mov", ".avi", ".mkv", ".webm"}:
        ctype = "upload_video"
    elif ext in {".mp3", ".m4a", ".wav", ".flac", ".ogg"}:
        ctype = "upload_audio"
    else:
        ctype = "upload_file"

    content = [{
        "type": ctype,
        "value": ",".join(file_ids),
        "key": args.file_key,
    }]

    raw = invoke_agent(args.app_id, args.api_key, content)
    result["elapsed_ms"] = round((time.time() - start) * 1000)
    if raw and "调用异常" not in raw and "API错误" not in raw \
              and "IllegalArgumentException" not in raw:
        result["output"]  = raw
        result["success"] = True
    else:
        result["status"] = "invoke_failed"
        result["output"] = raw or ""
    return result

# ── Excel ─────────────────────────────────────────────────────────────────────

COL_HEADERS = ["序号", "输入/文件组", "智能体输出", "状态", "耗时(ms)"]


def init_excel(path: str):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试结果"
    hf    = PatternFill("solid", fgColor="4472C4")
    hfont = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = hfont
        c.alignment = Alignment(horizontal="center")
    for col, w in enumerate([6, 50, 80, 14, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(path)


def append_row_excel(path: str, result: dict):
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill

    input_label = (result["input"] or
                   " | ".join(result["input_files"]))[:120]
    vals = [
        result["index"],
        input_label,
        result["output"],
        result["status"],
        result["elapsed_ms"],
    ]
    with _excel_lock:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        row = ws.max_row + 1
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v).alignment = Alignment(
                wrap_text=True)
        if result["status"] != "success":
            for col in range(1, len(COL_HEADERS) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(
                    "solid", fgColor="FFE0E0")
        wb.save(path)

# ── 主逻辑 ────────────────────────────────────────────────────────────────────

def build_text_tasks(args) -> list:
    """从 Excel/CSV 读取 prompt 列，生成文本任务列表"""
    path = args.input_file
    ext  = Path(path).suffix.lower()
    rows = []
    if ext in {".xlsx", ".xls"}:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(max_row=1))]
        col_idx = headers.index(args.input_col) if args.input_col in headers else 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[col_idx]
            if v is not None:
                rows.append(str(v))
    elif ext == ".csv":
        import csv
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for r in reader:
                v = r.get(args.input_col, "").strip()
                if v:
                    rows.append(v)
    else:
        # 纯文本，每行一个 prompt
        with open(path, encoding="utf-8") as f:
            rows = [l.strip() for l in f if l.strip()]

    limit = args.count if args.count > 0 else len(rows)
    return [{"index": i + 1, "prompt": p} for i, p in enumerate(rows[:limit])]


def build_file_tasks(args) -> list:
    """从目录扫描文件，按 group_size 分组，生成文件任务列表"""
    if args.image_dir:
        src_dir = args.image_dir
        valid   = IMAGE_EXTS
    else:
        src_dir = args.file_dir
        valid   = FILE_EXTS

    files = sorted([
        os.path.join(src_dir, f)
        for f in os.listdir(src_dir)
        if Path(f).suffix.lower() in valid
    ])
    group = args.group_size
    groups = [files[i:i + group] for i in range(0, len(files) - group + 1, group)]

    limit = args.count if args.count > 0 else len(groups)
    return [{"index": i + 1, "paths": g} for i, g in enumerate(groups[:limit])]


def print_progress(done: int, total: int, r: dict, start_time: float):
    avg = (time.time() - start_time) / done
    rem = avg * (total - done) / 60
    icon = "✓" if r["success"] else "✗"
    label = (r["input"] or " | ".join(r["input_files"]))[:40]
    print(f"[{done:04d}/{total}] {icon} {label} "
          f"({r['elapsed_ms']}ms) | 剩余≈{rem:.1f}分钟")


def main():
    parser = argparse.ArgumentParser(
        description="GLM 智能体批量测评工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--app-id", required=True,  help="智能体（应用）ID")
    parser.add_argument("--api-key", default=None,  help="智谱 API Key（默认读 ZHIPUAI_API_KEY）")

    # 文本模式
    parser.add_argument("--input-file", help="prompt 输入文件（xlsx/csv/txt）")
    parser.add_argument("--input-col",  default="input", help="prompt 列名（默认 input）")
    parser.add_argument("--input-key",  default="用户输入", help="对应智能体变量名（默认 用户输入）")

    # 图片模式
    parser.add_argument("--image-dir",       help="图片目录路径")
    parser.add_argument("--group-size", type=int, default=1,
                        help="每次调用的图片/文件数（默认 1）")

    # 文件/视频/音频模式
    parser.add_argument("--file-dir",        help="文件/视频/音频目录路径")
    parser.add_argument("--file-key",  default="图片", help="上传变量名（默认 图片）")

    # 通用上传参数
    parser.add_argument("--upload-unit-id",  help="上传组件 ID（图片/文件模式必填）")
    parser.add_argument("--file-type", type=int, default=4,
                        help="文件类型：1=excel 2=文档 3=音频 4=图片 5=视频（默认 4）")

    # 通用控制
    parser.add_argument("--concurrency", type=int, default=2, help="并发线程数（默认 2）")
    parser.add_argument("--count",       type=int, default=-1, help="测试数量上限（默认全部）")
    parser.add_argument("--output",      default="./output",   help="结果输出目录")
    parser.add_argument("--json-only",   action="store_true",  help="只输出 JSON 摘要")
    parser.add_argument("--list-vars",   action="store_true",  help="查询智能体输入变量后退出")

    args = parser.parse_args()

    # API Key
    args.api_key = args.api_key or os.environ.get("ZHIPUAI_API_KEY", "")
    if not args.api_key:
        print("错误：未提供 API Key，请用 --api-key 或设置 ZHIPUAI_API_KEY 环境变量",
              file=sys.stderr)
        sys.exit(1)

    # --list-vars 模式
    if args.list_vars:
        vars_ = list_variables(args.app_id, args.api_key)
        print(f"智能体 {args.app_id} 输入变量：")
        for v in vars_:
            print(f"  id={v['id']}  type={v['type']}  name={v['name']}")
        return

    # 构建任务列表
    if args.input_file:
        tasks    = build_text_tasks(args)
        mode     = "text"
    elif args.image_dir or args.file_dir:
        if not args.upload_unit_id:
            print("错误：图片/文件模式需要 --upload-unit-id，可用 --list-vars 查询",
                  file=sys.stderr)
            sys.exit(1)
        tasks = build_file_tasks(args)
        mode  = "file"
    else:
        parser.print_help()
        sys.exit(1)

    if not tasks:
        print("未找到任何测试任务，请检查输入路径/文件。", file=sys.stderr)
        sys.exit(1)

    # 初始化输出
    os.makedirs(args.output, exist_ok=True)
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path  = os.path.join(args.output, f"agent_test_{ts}.xlsx")
    init_excel(xlsx_path)

    if not args.json_only:
        print(f"▶ 智能体 {args.app_id}  任务数={len(tasks)}  并发={args.concurrency}")
        print(f"  结果 → {xlsx_path}")

    # 并发执行
    all_results = []
    done        = 0
    start_time  = time.time()
    process_fn  = process_text_task if mode == "text" else process_file_task

    with ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        futs = {ex.submit(process_fn, t, args): t for t in tasks}
        for fut in as_completed(futs):
            r = fut.result()
            all_results.append(r)
            append_row_excel(xlsx_path, r)
            done += 1
            if not args.json_only:
                print_progress(done, len(tasks), r, start_time)
            time.sleep(1.0)   # 简单限速，防止触发服务端限流

    # 统计
    elapsed  = time.time() - start_time
    succ_cnt = sum(1 for r in all_results if r["success"])
    fail_cnt = len(all_results) - succ_cnt
    status_counter = Counter(r["status"] for r in all_results)

    if not args.json_only:
        print(f"\n完成！耗时 {elapsed / 60:.1f} 分钟  "
              f"成功 {succ_cnt}/{len(all_results)}  失败 {fail_cnt}")
        print(f"状态分布: {dict(status_counter)}")

    # JSON 摘要输出
    summary = {
        "success":       succ_cnt == len(all_results),
        "app_id":        args.app_id,
        "total":         len(all_results),
        "success_count": succ_cnt,
        "fail_count":    fail_cnt,
        "elapsed_s":     round(elapsed, 1),
        "xlsx_path":     xlsx_path,
        "status_dist":   dict(status_counter),
        "results": sorted([
            {
                "index":       r["index"],
                "input_files": r["input_files"],
                "input":       r["input"][:200] if r["input"] else "",
                "success":     r["success"],
                "output":      r["output"][:500] if r["output"] else "",
                "elapsed_ms":  r["elapsed_ms"],
                "status":      r["status"],
            }
            for r in all_results
        ], key=lambda x: x["index"]),
    }

    print("\n--- JSON_RESULT_START ---")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("--- JSON_RESULT_END ---")


if __name__ == "__main__":
    main()
